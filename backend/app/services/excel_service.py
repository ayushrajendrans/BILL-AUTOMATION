
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
import io
from PIL import Image as PILImage

def create_excel_with_images(data_list):
    """
    Generates an Excel file in memory with embedded images.
    
    Args:
        data_list (list): List of dicts: {'date': str, 'club_name': str, 'image_bytes': bytes}
    
    Returns:
        bytes: The Excel file content.
    """
    # Create a workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Billing Data"
    
    # Define headers
    headers = ["Date", "Club Name", "Poster Image"]
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25  # Roughly 200px width

    row_height = 150 # Height in points, roughly 200px

    current_row = 2
    for item in data_list:
        # Write text data
        ws.cell(row=current_row, column=1, value=item.get('date', ''))
        ws.cell(row=current_row, column=2, value=item.get('club_name', ''))
        
        # Style text cells
        ws.cell(row=current_row, column=1).alignment = Alignment(vertical="center", horizontal="center")
        ws.cell(row=current_row, column=2).alignment = Alignment(vertical="center", horizontal="left")
        
        # Set row height
        ws.row_dimensions[current_row].height = row_height

        # Embed Image
        img_bytes = item.get('image_bytes')
        if img_bytes:
            try:
                # Open image with Pillow to resize
                pil_img = PILImage.open(io.BytesIO(img_bytes))
                
                # Resize to fit the cell (keeping aspect ratio)
                # Max height ~200px (150 points * 1.33)
                base_height = 190
                w_percent = (base_height / float(pil_img.size[1]))
                h_size = int((float(pil_img.size[0]) * float(w_percent)))
                pil_img = pil_img.resize((h_size, base_height), PILImage.Resampling.LANCZOS)
                
                # Convert back to bytes for Openpyxl
                img_stream = io.BytesIO()
                pil_img.save(img_stream, format='PNG')
                img_stream.seek(0)
                
                xl_img = OpenpyxlImage(img_stream)
                
                # Position image in Column C
                # Anchor to current row, column C (index 2)
                # Openpyxl anchors are 0-indexed for col/row in this context usually,
                # but let's use the cell coordinate
                cell_address = f"C{current_row}"
                ws.add_image(xl_img, cell_address)
                
            except Exception as e:
                print(f"Error processing image for row {current_row}: {e}")
                ws.cell(row=current_row, column=3, value="[Image Error]")

        current_row += 1

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.read()
